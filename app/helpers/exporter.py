import io
import csv
import json
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False


def _flatten_results(results: list[dict], bugs: list[str], dup_map: dict) -> list[dict]:
    rows = []
    for i, r in enumerate(results):
        dup_indices = dup_map.get(r["original_index"], [])
        dup_texts   = " | ".join(bugs[d][:80] for d in dup_indices)
        rows.append({
            "bug_number":     i + 1,
            "bug_text":       r.get("text", ""),
            "bug_type":       r.get("bug_type", ""),
            "severity":       r.get("severity", ""),
            "fix_time":       r.get("fix_time", ""),
            "bt_confidence":  r.get("bt_conf", 0),
            "sv_confidence":  r.get("sv_conf", 0),
            "ft_confidence":  r.get("ft_conf", 0),
            "is_uncertain":   r.get("is_uncertain", False),
            "duplicate_count": len(dup_indices),
            "duplicate_texts": dup_texts,
            "timestamp":      datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        })
    return rows


def export_csv(results: list[dict], bugs: list[str], dup_map: dict) -> bytes:
    rows   = _flatten_results(results, bugs, dup_map)
    buf    = io.StringIO()
    if not rows:
        return b""
    writer = csv.DictWriter(buf, fieldnames=rows[0].keys())
    writer.writeheader()
    writer.writerows(rows)
    return buf.getvalue().encode("utf-8")


def export_xlsx(results: list[dict], bugs: list[str], dup_map: dict) -> bytes:
    if not XLSX_AVAILABLE:
        raise ImportError("openpyxl is not installed. Run: pip install openpyxl")

    rows = _flatten_results(results, bugs, dup_map)
    wb   = openpyxl.Workbook()
    ws   = wb.active
    ws.title = "Bug Classification Report"

    header_fill = PatternFill("solid", fgColor="7C5CFC")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    headers     = list(rows[0].keys()) if rows else []

    for col, h in enumerate(headers, 1):
        cell              = ws.cell(row=1, column=col, value=h.replace("_", " ").title())
        cell.fill         = header_fill
        cell.font         = header_font
        cell.alignment    = Alignment(horizontal="center", vertical="center")

    sev_colors = {
        "critical":  "F87171",
        "major":     "FB923C",
        "minor":     "34D399",
        "uncertain": "A78BFA",
    }

    for row_idx, row in enumerate(rows, 2):
        for col_idx, key in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row[key])
            if key == "severity":
                color = sev_colors.get(str(row[key]).lower(), "FFFFFF")
                cell.fill = PatternFill("solid", fgColor=color)
            if key == "is_uncertain" and row[key]:
                cell.fill = PatternFill("solid", fgColor="FEF08A")

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def export_json(results: list[dict], bugs: list[str], dup_map: dict) -> bytes:
    rows = _flatten_results(results, bugs, dup_map)
    return json.dumps(rows, indent=2).encode("utf-8")
